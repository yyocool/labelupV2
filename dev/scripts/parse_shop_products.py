#!/usr/bin/env python3
"""Parse LabelUp product Excel and import into shop_products."""
from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

COL = {
    "group": 1,
    "sku": 3,
    "product_no": 4,
    "art_no": 5,
    "barcode_no": 6,
    "material_name": 7,
    "barcode": 8,
    "box_barcode": 9,
    "paper_size": 10,
    "labels_per_sheet": 11,
    "std_size": 12,
    "spec_mm": 13,
    "pack_size": 14,
    "box_size": 15,
    "sheets_per_pack": 16,
    "qty_per_box": 17,
    "material": 18,
    "weight": 19,
    "thickness": 20,
    "origin": 21,
    "etc": 22,
    "product_name": 24,
    "price": 25,
    "sale_price": 26,
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def parse_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    current_group = None

    for r in range(4, ws.max_row + 1):
        group = clean(ws.cell(r, COL["group"]).value)
        if group:
            current_group = str(group)

        sku = clean(ws.cell(r, COL["sku"]).value)
        name = clean(ws.cell(r, COL["product_name"]).value)
        if not sku and not name:
            continue

        sku = str(sku).strip() if sku else ""
        name = str(name).strip() if name else ""
        if not sku and name:
            sku = re.sub(r"[^\w\-]+", "-", name.split("/")[0]).strip("-")[:80]
        if not sku:
            continue
        if not name:
            material_name = clean(ws.cell(r, COL["material_name"]).value) or "라벨"
            labels = clean(ws.cell(r, COL["labels_per_sheet"]).value)
            sheets = clean(ws.cell(r, COL["sheets_per_pack"]).value)
            parts = [sku, str(material_name)]
            if labels:
                parts.append(f"{labels}칸")
            if sheets:
                parts.append(f"{sheets}매")
            name = "/".join(parts)

        rows.append(
            {
                "row": r,
                "group": current_group or "라벨지",
                "sku": sku,
                "name": name,
                "product_no": clean(ws.cell(r, COL["product_no"]).value),
                "art_no": clean(ws.cell(r, COL["art_no"]).value),
                "barcode_no": clean(ws.cell(r, COL["barcode_no"]).value),
                "material_name": clean(ws.cell(r, COL["material_name"]).value),
                "barcode": clean(ws.cell(r, COL["barcode"]).value),
                "box_barcode": clean(ws.cell(r, COL["box_barcode"]).value),
                "paper_size": clean(ws.cell(r, COL["paper_size"]).value),
                "labels_per_sheet": clean(ws.cell(r, COL["labels_per_sheet"]).value),
                "std_size": clean(ws.cell(r, COL["std_size"]).value),
                "spec_mm": clean(ws.cell(r, COL["spec_mm"]).value),
                "pack_size": clean(ws.cell(r, COL["pack_size"]).value),
                "box_size": clean(ws.cell(r, COL["box_size"]).value),
                "sheets_per_pack": clean(ws.cell(r, COL["sheets_per_pack"]).value),
                "qty_per_box": clean(ws.cell(r, COL["qty_per_box"]).value),
                "material": clean(ws.cell(r, COL["material"]).value),
                "weight": clean(ws.cell(r, COL["weight"]).value),
                "thickness": clean(ws.cell(r, COL["thickness"]).value),
                "origin": clean(ws.cell(r, COL["origin"]).value),
                "etc": clean(ws.cell(r, COL["etc"]).value),
                "price": clean(ws.cell(r, COL["price"]).value),
                "sale_price": clean(ws.cell(r, COL["sale_price"]).value),
                "stock_qty": 100,
                "status": "active",
            }
        )

    return rows


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    default_xlsx = root / "storage" / "imports" / "labelup-products-schedule.xlsx"
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else default_xlsx
    out_json = root / "storage" / "imports" / "products_import.json"

    if not xlsx.exists():
        print(f"Excel not found: {xlsx}", file=sys.stderr)
        return 1

    rows = parse_excel(xlsx)
    out_json.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Parsed {len(rows)} products -> {out_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
