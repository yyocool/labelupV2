# -*- coding: utf-8 -*-
"""
storyboard/ 각 페이지의 "내용"(메타/영역/기능/UX)을 파싱해
전체 프로젝트 기능명세서(.xlsx)를 생성한다.

- 관리자형(02-*): $adminMeta / $adminZones / $adminUx PHP 배열
- 프론트형(01-*): sb-front-meta-grid / sb-front-spec 테이블(HTML) 또는 zone-data 프래그먼트
- 스텁(sb-page--placeholder): 미작성(작업예정)

사용법: python tools/build_function_spec.py
출력:   기능명세서_YYMMDD.xlsx (워크스페이스 루트)
"""
import os
import re
import html
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SB_DIR = os.path.join(ROOT, "project", "storyboard")
FRAG_DIR = os.path.join(SB_DIR, "_fragments")

CODE_RE = re.compile(r"^\d+(?:-\d+)*$")


# ---------------------------------------------------------------- helpers
def php_unescape(s):
    """PHP 작은따옴표 문자열 escape 해제."""
    return s.replace("\\'", "'").replace("\\\\", "\\")


def strip_html(s, code=""):
    if s is None:
        return ""
    # PHP echo 토큰을 실제 코드로 치환/제거
    s = re.sub(r"<\?=?\s*e\(\$(?:menuCode|adminCode|code)\)\s*;?\s*\?>", code, s)
    s = re.sub(r"<\?php.*?\?>", "", s, flags=re.S)
    s = re.sub(r"<\?=.*?\?>", "", s, flags=re.S)
    s = re.sub(r"<br\s*/?>", " · ", s, flags=re.I)
    s = re.sub(r"<[^>]+>", "", s)          # 나머지 태그 제거
    s = html.unescape(s)
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # 앞뒤 잔여 구분자 정리
    s = s.strip(" ·|,")
    return s.strip()


def read(path):
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def parse_php_kv(entry):
    """array('k'=>'v', ...) 내부에서 key=>'value' 추출."""
    out = {}
    for k, v in re.findall(r"'(\w+)'\s*=>\s*'((?:[^'\\]|\\.)*)'", entry):
        out[k] = php_unescape(v)
    return out


# ---------------------------------------------------------------- parse meta
def parse_meta(text, code):
    meta = {}
    m = re.search(r"\$adminMeta\s*=\s*array\((.*?)\);", text, re.S)
    if m:
        for entry in re.findall(r"array\((.*?)\)", m.group(1), re.S):
            kv = parse_php_kv(entry)
            if "dt" in kv:
                meta.setdefault(strip_html(kv.get("dt"), code),
                                strip_html(kv.get("dd", ""), code))
    for dt, dd in re.findall(r"<dt>(.*?)</dt>\s*<dd>(.*?)</dd>", text, re.S):
        meta.setdefault(strip_html(dt, code), strip_html(dd, code))
    return meta


# ---------------------------------------------------------------- parse zones
KIND_LABEL = {
    "nav": "Nav", "ui": "UI", "cta": "CTA", "layout": "Layout",
    "form": "Form", "data": "Data",
}


def parse_zones(text, code):
    zones = []  # dict(id, kind, block, el, link)

    # 1) 관리자형 PHP 배열
    m = re.search(r"\$adminZones\s*=\s*array\((.*?)\n\);", text, re.S)
    if not m:
        m = re.search(r"\$adminZones\s*=\s*array\((.*?)\);", text, re.S)
    if m:
        for entry in re.findall(r"array\((.*?)\)", m.group(1), re.S):
            kv = parse_php_kv(entry)
            if kv.get("id") or kv.get("block"):
                zones.append({
                    "id": strip_html(kv.get("id", ""), code),
                    "kind": KIND_LABEL.get(kv.get("kind", "").lower(),
                                           kv.get("kind", "")),
                    "block": strip_html(kv.get("block", ""), code),
                    "el": strip_html(kv.get("el", ""), code),
                    "link": strip_html(kv.get("link", ""), code),
                })

    # 2) 프론트형 HTML 스펙 테이블(영역 ID 5열)
    if not zones:
        for sec in re.findall(r'<section class="sb-front-spec">(.*?)</section>',
                              text, re.S):
            head = re.search(r"<h3>(.*?)</h3>", sec, re.S)
            head_txt = strip_html(head.group(1), code) if head else ""
            thead = re.search(r"<thead>(.*?)</thead>", sec, re.S)
            is_zone = ("영역" in (thead.group(1) if thead else "")) or \
                      ("구조" in head_txt)
            if not is_zone:
                continue
            body = re.search(r"<tbody>(.*?)</tbody>", sec, re.S)
            if not body:
                continue
            for row in re.findall(r"<tr>(.*?)</tr>", body.group(1), re.S):
                cells = re.findall(r"<td.*?>(.*?)</td>", row, re.S)
                if len(cells) < 4:
                    continue
                kind_raw = ""
                km = re.search(r"tag--(\w+)", cells[1])
                if km:
                    kind_raw = km.group(1)
                zones.append({
                    "id": strip_html(cells[0], code),
                    "kind": KIND_LABEL.get(kind_raw.lower(), strip_html(cells[1], code)),
                    "block": strip_html(cells[2], code),
                    "el": strip_html(cells[3], code),
                    "link": strip_html(cells[4], code) if len(cells) > 4 else "",
                })

    # 3) zone-data 프래그먼트 fallback
    if not zones:
        frag = os.path.join(FRAG_DIR, "zone-data-%s.php" % code)
        if os.path.exists(frag):
            ftext = read(frag)
            for zid, entry in re.findall(
                    r"'([A-Z]-\d+)'\s*=>\s*array\((.*?)\)\s*,", ftext, re.S):
                kv = parse_php_kv(entry)
                zones.append({
                    "id": zid,
                    "kind": KIND_LABEL.get(kv.get("typeKey", "").lower(),
                                           kv.get("type", "")),
                    "block": strip_html(kv.get("block", ""), code),
                    "el": strip_html(kv.get("elements", ""), code),
                    "link": strip_html(kv.get("menu", ""), code),
                })
    return zones


# ---------------------------------------------------------------- parse ux
def parse_ux(text, code):
    ux = []  # (item, desc)
    m = re.search(r"\$adminUx\s*=\s*array\((.*?)\n\);", text, re.S)
    if not m:
        m = re.search(r"\$adminUx\s*=\s*array\((.*?)\);", text, re.S)
    if m:
        for entry in re.findall(r"array\((.*?)\)", m.group(1), re.S):
            kv = parse_php_kv(entry)
            if kv.get("item"):
                ux.append((strip_html(kv.get("item"), code),
                           strip_html(kv.get("desc", ""), code)))

    if not ux:
        for sec in re.findall(r'<section class="sb-front-spec">(.*?)</section>',
                              text, re.S):
            head = re.search(r"<h3>(.*?)</h3>", sec, re.S)
            head_txt = strip_html(head.group(1), code) if head else ""
            thead = re.search(r"<thead>(.*?)</thead>", sec, re.S)
            th_txt = thead.group(1) if thead else ""
            is_ux = ("UX" in head_txt or "인터랙션" in head_txt or "기능" in head_txt) \
                and "영역" not in th_txt
            if not is_ux:
                continue
            body = re.search(r"<tbody>(.*?)</tbody>", sec, re.S)
            if not body:
                continue
            for row in re.findall(r"<tr>(.*?)</tr>", body.group(1), re.S):
                cells = re.findall(r"<td.*?>(.*?)</td>", row, re.S)
                if len(cells) >= 2:
                    ux.append((strip_html(cells[0], code),
                               strip_html(cells[1], code)))
    return ux


# ---------------------------------------------------------------- parse page
def parse_title(text):
    m = re.search(r"스토리보드\s*[:：]\s*(.+)", text)
    if m:
        return m.group(1).strip()
    m = re.search(r"\$adminTitle\s*=\s*'((?:[^'\\]|\\.)*)'", text)
    if m:
        return php_unescape(m.group(1)).strip()
    return ""


def parse_page(path):
    code = os.path.splitext(os.path.basename(path))[0]
    text = read(path)
    is_stub = "sb-page--placeholder" in text
    has_hifi = ("hifi-wireframe-body" in text) or ("sb-wf--hifi" in text)

    title = parse_title(text) or code
    meta = parse_meta(text, code)
    zones = [] if is_stub else parse_zones(text, code)
    ux = [] if is_stub else parse_ux(text, code)

    if is_stub:
        status = "미작성(작업예정)"
    elif has_hifi:
        status = "완료(와이어프레임)"
    elif zones or ux:
        status = "명세 작성"
    else:
        status = "기타/개요"

    artifacts = []
    if zones or ux:
        artifacts.append("기능명세")
    if has_hifi:
        artifacts.append("Hi-Fi 와이어프레임")
    if is_stub:
        artifacts.append("스텁")

    return {
        "code": code,
        "title": title,
        "meta": meta,
        "zones": zones,
        "ux": ux,
        "status": status,
        "artifacts": " · ".join(artifacts) if artifacts else "-",
        "is_stub": is_stub,
    }


# ---------------------------------------------------------------- collect
def collect():
    pages = {}
    for name in os.listdir(SB_DIR):
        if not name.endswith(".php"):
            continue
        if name == "build_stubs.php":
            continue
        code = name[:-4]
        if not CODE_RE.match(code):
            continue
        pages[code] = parse_page(os.path.join(SB_DIR, name))
    return pages


def code_sort_key(code):
    return [int(x) for x in code.split("-")]


def breadcrumb(code, title_map):
    parts = code.split("-")
    crumbs = []
    for i in range(1, len(parts) + 1):
        pc = "-".join(parts[:i])
        crumbs.append(title_map.get(pc, pc))
    return crumbs


# ---------------------------------------------------------------- meta pickers
def pick(meta, *keys):
    for k in keys:
        for mk, mv in meta.items():
            if k in mk:
                return mv
    return ""


# ---------------------------------------------------------------- styling
HEAD_FILL = PatternFill("solid", fgColor="4F46E5")
SUB_FILL = PatternFill("solid", fgColor="EEF2FF")
STUB_FILL = PatternFill("solid", fgColor="FFF7ED")
DONE_FILL = PatternFill("solid", fgColor="ECFDF5")
SPEC_FILL = PatternFill("solid", fgColor="EFF6FF")
TITLE_FONT = Font(name="맑은 고딕", size=18, bold=True, color="1E293B")
HEAD_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="맑은 고딕", size=10, color="1E293B")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def apply_body(ws, first_row, ncols):
    for r in range(first_row, ws.max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER


# ---------------------------------------------------------------- build xlsx
def build():
    pages = collect()
    title_map = {c: p["title"] for c, p in pages.items()}
    codes = sorted(pages.keys(), key=code_sort_key)
    today = datetime.date.today()

    wb = Workbook()

    # ===== Sheet 1: 개요 =====
    ws = wb.active
    ws.title = "개요"
    ws["A1"] = "Label-UP 프로젝트 기능명세서"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = "스토리보드(storyboard/) 각 화면 내용 기반 · 자동 생성"
    ws["A2"].font = Font(name="맑은 고딕", size=10, color="64748B")
    ws.merge_cells("A2:D2")
    ws["A3"] = "생성일: %s" % today.strftime("%Y-%m-%d")
    ws["A3"].font = Font(name="맑은 고딕", size=10, color="64748B")

    total = len(pages)
    stub = sum(1 for p in pages.values() if p["is_stub"])
    hifi = sum(1 for p in pages.values() if "와이어프레임" in p["status"])
    spec = sum(1 for p in pages.values() if p["status"] == "명세 작성")
    other = total - stub - hifi - spec
    feat_cnt = sum(len(p["zones"]) + len(p["ux"]) for p in pages.values())

    summary = [
        ("항목", "값"),
        ("전체 화면(페이지) 수", total),
        ("완료(Hi-Fi 와이어프레임)", hifi),
        ("명세 작성", spec),
        ("개요/기타", other),
        ("미작성(작업예정 · 스텁)", stub),
        ("도출된 기능/영역 항목 수", feat_cnt),
    ]
    r0 = 5
    for i, (a, b) in enumerate(summary):
        ws.cell(row=r0 + i, column=1, value=a)
        ws.cell(row=r0 + i, column=2, value=b)
    style_header(ws, r0, 2)
    for r in range(r0 + 1, r0 + len(summary)):
        for c in (1, 2):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    # 대분류별 집계
    cat_r = r0 + len(summary) + 2
    ws.cell(row=cat_r, column=1, value="대분류").value = "대분류(코드)"
    ws.cell(row=cat_r, column=2, value="화면 수")
    ws.cell(row=cat_r, column=3, value="미작성")
    ws.cell(row=cat_r, column=4, value="완료/명세")
    style_header(ws, cat_r, 4)
    top_codes = sorted({c.split("-")[0] for c in codes})
    rr = cat_r + 1
    for tc in top_codes:
        members = [c for c in codes if c.split("-")[0] == tc]
        m_stub = sum(1 for c in members if pages[c]["is_stub"])
        ws.cell(row=rr, column=1,
                value="%s %s" % (tc, title_map.get(tc, "")))
        ws.cell(row=rr, column=2, value=len(members))
        ws.cell(row=rr, column=3, value=m_stub)
        ws.cell(row=rr, column=4, value=len(members) - m_stub)
        for c in range(1, 5):
            cell = ws.cell(row=rr, column=c)
            cell.font = CELL_FONT
            cell.border = BORDER
        rr += 1

    note_r = rr + 1
    ws.cell(row=note_r, column=1,
            value=("※ 상태 정의 — 완료: Hi-Fi 와이어프레임까지 산출 / "
                   "명세 작성: 화면 구성·기능 정의 완료 / "
                   "미작성: 스텁(작업 예정). 개발상태는 전 화면 '예정(미착수)'."))
    ws.cell(row=note_r, column=1).font = Font(name="맑은 고딕", size=9,
                                              color="94A3B8")
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=6)
    for w, col in zip([34, 14, 12, 14], "ABCD"):
        ws.column_dimensions[col].width = w

    # ===== Sheet 2: 화면목록 =====
    ws2 = wb.create_sheet("화면목록")
    headers2 = ["대분류", "화면코드", "뎁스", "메뉴 경로", "화면명", "화면 목적",
                "예상 URL", "접근 권한", "연관 데이터/테이블", "주요 기능(요약)",
                "산출물", "스토리보드 상태", "개발 상태", "비고"]
    ws2.append(headers2)
    style_header(ws2, 1, len(headers2))
    for code in codes:
        p = pages[code]
        m = p["meta"]
        crumbs = breadcrumb(code, title_map)
        top = "%s %s" % (code.split("-")[0], title_map.get(code.split("-")[0], ""))
        purpose = pick(m, "목적")
        url = pick(m, "URL")
        auth = pick(m, "권한")
        tables = pick(m, "연관", "테이블", "데이터")
        blocks = [z["block"] for z in p["zones"] if z["block"]]
        summary_feat = ", ".join(blocks[:8])
        if len(blocks) > 8:
            summary_feat += " 외 %d건" % (len(blocks) - 8)
        if not summary_feat and p["ux"]:
            summary_feat = ", ".join(u[0] for u in p["ux"][:6])
        ws2.append([
            top, code, len(code.split("-")), " › ".join(crumbs), p["title"],
            purpose, url, auth, tables, summary_feat, p["artifacts"],
            p["status"], "예정(미착수)",
            "스텁 — 상세 명세 필요" if p["is_stub"] else "",
        ])
        # 상태 색상
        fill = STUB_FILL if p["is_stub"] else (
            DONE_FILL if "완료" in p["status"] else (
                SPEC_FILL if p["status"] == "명세 작성" else None))
        if fill:
            ws2.cell(row=ws2.max_row, column=12).fill = fill
    apply_body(ws2, 2, len(headers2))
    ws2.freeze_panes = "A2"
    widths2 = [16, 14, 6, 30, 22, 34, 20, 18, 26, 40, 20, 18, 14, 22]
    for i, w in enumerate(widths2):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    # ===== Sheet 3: 기능명세(상세) =====
    ws3 = wb.create_sheet("기능명세(상세)")
    headers3 = ["화면코드", "화면명", "영역/기능 ID", "구분", "기능/블록명",
                "상세 내용(포함 요소)", "연결 화면", "개발 상태"]
    ws3.append(headers3)
    style_header(ws3, 1, len(headers3))
    for code in codes:
        p = pages[code]
        if not p["zones"] and not p["ux"]:
            continue
        for z in p["zones"]:
            ws3.append([code, p["title"], z["id"], z["kind"] or "-",
                        z["block"], z["el"], z["link"] or "-", "예정"])
        for item, desc in p["ux"]:
            ws3.append([code, p["title"], "-", "인터랙션/UX",
                        item, desc, "-", "예정"])
    apply_body(ws3, 2, len(headers3))
    ws3.freeze_panes = "A2"
    widths3 = [12, 22, 14, 12, 26, 60, 14, 10]
    for i, w in enumerate(widths3):
        ws3.column_dimensions[get_column_letter(i + 1)].width = w

    # ===== Sheet 4: 미작성(작업예정) 목록 =====
    ws4 = wb.create_sheet("작업예정(미작성)")
    headers4 = ["대분류", "화면코드", "메뉴 경로", "화면명", "개발 상태", "비고"]
    ws4.append(headers4)
    style_header(ws4, 1, len(headers4))
    for code in codes:
        p = pages[code]
        if not p["is_stub"]:
            continue
        crumbs = breadcrumb(code, title_map)
        top = "%s %s" % (code.split("-")[0], title_map.get(code.split("-")[0], ""))
        ws4.append([top, code, " › ".join(crumbs), p["title"],
                    "예정(미착수)", "스토리보드 명세 작성 필요"])
    apply_body(ws4, 2, len(headers4))
    ws4.freeze_panes = "A2"
    for i, w in enumerate([16, 14, 34, 24, 14, 26]):
        ws4.column_dimensions[get_column_letter(i + 1)].width = w

    out = os.path.join(ROOT, "기능명세서_%s.xlsx" % today.strftime("%y%m%d"))
    wb.save(out)
    print("SAVED:", out)
    print("pages=%d stub=%d hifi=%d spec=%d other=%d features=%d"
          % (total, stub, hifi, spec, other, feat_cnt))


if __name__ == "__main__":
    build()
